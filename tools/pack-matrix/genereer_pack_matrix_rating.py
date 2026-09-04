#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genereer_Pack_Matrix_Rating
===========================
Zelfde matrix als Genereer_Pack_Matrix, maar de rijen en kolommen staan
gesorteerd op overall in plaats van op naam, en achter elke naam staat die
overall tussen haakjes: "Petar (89)".

Daarvoor is de *volledige* pack-historie nodig, niet de compacte regels:

    GET api/packs                 -> Input.txt   (dit script)
    GET api/packs?compact=true                   (het andere script)

De compacte regels dragen de overall van de kaarten wel ("Ton (74)"), maar niet
die van de packer, geen speler-id om op te sleutelen en geen icoon-vlag. De
volledige respons bevat per kaart het hele subject. Sla die respons op als
Input.txt naast dit script.

Let op: de overall in de respons is de overall van *nu*, niet die op het moment
van pakken - de API leest elk subject van de levende kaartenpool. Dat is ook de
enige overall waarop een as te sorteren valt: iemand komt in tientallen packs
voor, en een as heeft een getal per persoon nodig.

Dit script wordt normaal gestart via Genereer_Pack_Matrix_Rating.bat, die in
dezelfde map moet staan als Input.txt.
"""

import json
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FOUT: de module 'openpyxl' is niet gevonden.")
    print("Installeer deze met:  pip install openpyxl")
    sys.exit(1)


def script_dir():
    return os.path.dirname(os.path.abspath(__file__))


def field(obj, name, default=None):
    """Haalt een veld hoofdletterongevoelig op, zodat zowel camelCase
    ("claimedAt") als PascalCase ("ClaimedAt") werkt - welke van de twee de API
    teruggeeft hangt af van de serialisatie-instellingen, niet van dit script."""
    if not isinstance(obj, dict):
        return default
    if name in obj:
        return obj[name]
    lowered = name.lower()
    for key, value in obj.items():
        if key.lower() == lowered:
            return value
    return default


def load_packs(path):
    """Geeft de lijst met packs terug. Accepteert zowel het hele
    PackHistory-object als een kale lijst met packs."""
    with open(path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    if isinstance(data, dict):
        packs = field(data, "packs")
        if packs is None:
            raise ValueError(
                "Input.txt bevat een JSON-object zonder 'packs'. Verwacht wordt de "
                "respons van GET api/packs (zonder compact=true)."
            )
    elif isinstance(data, list):
        packs = data
        if packs and isinstance(packs[0], str):
            raise ValueError(
                "Input.txt bevat de compacte regels (GET api/packs?compact=true). "
                "Dit script heeft de volledige respons nodig: GET api/packs. "
                "Gebruik anders Genereer_Pack_Matrix.bat."
            )
    else:
        raise ValueError("Input.txt moet de respons van GET api/packs bevatten.")

    if not isinstance(packs, list):
        raise ValueError("'packs' is geen lijst.")
    return packs


def first_name(name):
    """De eerste naam, net zoals de kaart en de compacte regel hem drukken."""
    parts = name.split() if name else []
    return parts[0] if parts else name


def parse_packs(packs):
    """Geeft (counts, people, stats, errors) terug.

    counts: dict[(packer_id, gepackte_id)] -> aantal keer
    people: dict[id] -> {"name", "overall", "rating", "icon", "games"}
            overall is None voor iemand die zelf geen kaart heeft (te weinig
            wedstrijden) maar wel packs opent.
    stats:  losse tellers voor de terugkoppeling op het scherm
    errors: packs die niet te lezen waren
    """
    counts = {}
    people = {}
    errors = []
    total_cards = 0
    good_packs = 0

    def ensure(person_id, name):
        entry = people.get(person_id)
        if entry is None:
            entry = {
                "name": name,
                "overall": None,
                "rating": None,
                "icon": False,
                "games": None,
            }
            people[person_id] = entry
        elif name and not entry["name"]:
            entry["name"] = name
        return entry

    for pack in packs:
        packer_id = field(pack, "collectorId")
        packer_name = field(pack, "collectorName")
        cards = field(pack, "cards")

        if not packer_id or not isinstance(cards, list):
            errors.append(pack)
            continue

        ensure(packer_id, packer_name or packer_id)

        broken = False
        for card in cards:
            subject = field(card, "subject")
            packed_id = field(subject, "id")
            if not packed_id:
                # Eenmaal per pack tellen, ook als er meer kaarten in stuk zijn -
                # anders zegt de teller straks dat er meer packs kapot waren dan er
                # packs zijn.
                broken = True
                continue

            entry = ensure(packed_id, field(subject, "name") or packed_id)
            # Het subject komt van de levende pool, dus elk voorkomen draagt
            # dezelfde overall; de laatste overschrijft de vorige zonder verschil.
            entry["overall"] = field(subject, "overall")
            entry["rating"] = field(subject, "visibleRating")
            entry["icon"] = bool(field(subject, "isIcon", False))
            entry["games"] = field(subject, "numberOfGames")

            key = (packer_id, packed_id)
            counts[key] = counts.get(key, 0) + 1
            total_cards += 1

        if broken:
            errors.append(pack)
        else:
            good_packs += 1

    stats = {"packs": good_packs, "cards": total_cards}
    return counts, people, stats, errors


def sort_people(people):
    """De as: hoogste overall eerst, dan hoogste rating, dan naam. Wie geen
    kaart heeft (overall None) staat achteraan, alfabetisch - anders zou een
    ontbrekend getal als een 0 sorteren en tussen de laagste kaarten belanden."""
    def key(item):
        person_id, entry = item
        overall = entry["overall"]
        rating = entry["rating"]
        return (
            0 if overall is not None else 1,
            -(overall or 0),
            -(rating or 0),
            (entry["name"] or "").lower(),
            person_id,
        )

    return [person_id for person_id, _ in sorted(people.items(), key=key)]


def label(entry):
    """"Petar (89)", of alleen de naam als die persoon geen kaart heeft."""
    name = first_name(entry["name"])
    overall = entry["overall"]
    return f"{name} ({overall})" if overall is not None else name


def build_workbook(counts, people, order):
    n = len(order)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pack matrix (rating)"

    bold = Font(bold=True)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    diag_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    icon_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    slanted = Alignment(horizontal="center", vertical="bottom", textRotation=60)
    stacked = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    def head(row, col, value, alignment=center):
        c = ws.cell(row=row, column=col, value=value)
        c.font = header_font
        c.fill = header_fill
        c.alignment = alignment
        c.border = border
        return c

    def total_cell(row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = bold
        c.fill = total_fill
        c.alignment = center
        c.border = border
        return c

    # Twee totaalkolommen, want ze zeggen verschillende dingen: vier keer dezelfde
    # kaart is vier kaarten maar een unieke.
    #
    # Ze staan vooraan in plaats van achteraan, tegen de gewoonte in. Met zevenendertig
    # spelers is de matrix veertig kolommen breed, en Excel kan alleen aan de linkerkant
    # bevriezen - achteraan zouden juist de twee getallen die je leest van het scherm af
    # staan. Vooraan reizen ze mee met de naam. De twee totaalrijen blijven wel onderaan:
    # veertig rijen passen bijna op een scherm en veertig kolommen bij lange na niet, dus
    # de assen zijn scheef omdat het blad zelf scheef is.
    total_col = 2
    unique_col = 3
    first_col = 4

    # Rij 1: kolomkoppen (de gepackten), op overall van hoog naar laag. Schuin,
    # want "Petar (89)" past niet in een kolom van tien tekens breed.
    head(1, 1, "Packer \\ Gepackte")
    # Deze twee gewikkeld in plaats van schuin: op 60 graden zou "Totale unieke
    # kaarten gepackt" de hele kopregel twee keer zo hoog maken voor twee kolommen.
    head(1, total_col, "Totale kaarten gepackt", alignment=stacked)
    head(1, unique_col, "Totale unieke kaarten gepackt", alignment=stacked)
    for j, person_id in enumerate(order, start=first_col):
        entry = people[person_id]
        c = head(1, j, label(entry), alignment=slanted)
        if entry["icon"]:
            c.fill = icon_fill

    # Datarijen (de packers), in dezelfde volgorde als de kolommen.
    col_totals = [0] * n
    col_uniques = [0] * n
    for i, packer_id in enumerate(order):
        row = i + 2
        entry = people[packer_id]
        c = head(row, 1, label(entry))
        if entry["icon"]:
            c.fill = icon_fill

        row_total = 0
        row_unique = 0
        for j, packed_id in enumerate(order):
            value = counts.get((packer_id, packed_id), 0)
            cell = ws.cell(row=row, column=j + first_col, value=value if value else None)
            cell.alignment = center
            cell.border = border
            if packer_id == packed_id:
                cell.fill = diag_fill
            row_total += value
            col_totals[j] += value
            if value:
                row_unique += 1
                col_uniques[j] += 1

        total_cell(row, total_col, row_total)
        total_cell(row, unique_col, row_unique)

    # Totaalrij: hoe vaak iemand zelf gepackt is, en door hoeveel verschillende
    # packers - de spiegel van de twee totaalkolommen.
    total_row = n + 2
    unique_row = n + 3
    head(total_row, 1, "Totaal gepackt")
    head(unique_row, 1, "Door unieke packers")
    for j in range(n):
        total_cell(total_row, j + first_col, col_totals[j])
        total_cell(unique_row, j + first_col, col_uniques[j])

    # De hoek: alle kaarten, en alle paren packer-gepackte die ooit voorkwamen. De
    # twee overgebleven hoekvakjes blijven leeg - daar zou alleen het getal van de
    # tegenoverliggende hoek nog een keer staan.
    total_cell(total_row, total_col, sum(col_totals))
    total_cell(unique_row, unique_col, sum(col_uniques))
    total_cell(total_row, unique_col, None)
    total_cell(unique_row, total_col, None)

    ws.column_dimensions["A"].width = 18
    for j in (total_col, unique_col):
        ws.column_dimensions[get_column_letter(j)].width = 14
    for j in range(first_col, first_col + n):
        ws.column_dimensions[get_column_letter(j)].width = 11
    # Naam en beide totalen blijven staan tijdens het scrollen; dat is de hele reden
    # dat ze vooraan staan.
    ws.freeze_panes = "D2"
    ws.row_dimensions[1].height = 95

    return wb


def build_summary_sheet(wb, counts, people):
    """Platte lijst, gesorteerd op aantal - met de overalls erbij, zodat te zien
    is of de dure kaarten ook echt bij weinig mensen vandaan komen."""
    ws = wb.create_sheet("Overzicht (lijst)")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Packer",
        "Overall packer",
        "Gepackte",
        "Overall gepackte",
        "Rating gepackte",
        "Icoon",
        "Aantal keer gepackt",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill

    rows = sorted(
        counts.items(),
        key=lambda kv: (
            -kv[1],
            -(people[kv[0][1]]["overall"] or 0),
            (people[kv[0][0]]["name"] or "").lower(),
            (people[kv[0][1]]["name"] or "").lower(),
        ),
    )
    for i, ((packer_id, packed_id), aantal) in enumerate(rows, start=2):
        packer = people[packer_id]
        packed = people[packed_id]
        ws.cell(row=i, column=1, value=first_name(packer["name"]))
        ws.cell(row=i, column=2, value=packer["overall"])
        ws.cell(row=i, column=3, value=first_name(packed["name"]))
        ws.cell(row=i, column=4, value=packed["overall"])
        ws.cell(row=i, column=5, value=packed["rating"])
        ws.cell(row=i, column=6, value="ja" if packed["icon"] else "nee")
        ws.cell(row=i, column=7, value=aantal)

    for column, width in zip("ABCDEFG", (16, 14, 16, 18, 17, 8, 20)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(rows) + 1}" if rows else "A1:G1"


def build_people_sheet(wb, counts, people, order):
    """Een regel per persoon, in dezelfde volgorde als de matrix: wat iemand
    packte, hoe vaak hij zelf gepackt werd, en door hoeveel verschillende packers."""
    ws = wb.create_sheet("Spelers (op overall)")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Naam",
        "Overall",
        "Rating",
        "Wedstrijden",
        "Icoon",
        "Kaarten gepackt",
        "Keer gepackt",
        "Door hoeveel packers",
    ]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = header_font
        c.fill = header_fill

    given = {}
    received = {}
    owners = {}
    for (packer_id, packed_id), aantal in counts.items():
        given[packer_id] = given.get(packer_id, 0) + aantal
        received[packed_id] = received.get(packed_id, 0) + aantal
        owners[packed_id] = owners.get(packed_id, 0) + 1

    for i, person_id in enumerate(order, start=2):
        entry = people[person_id]
        ws.cell(row=i, column=1, value=first_name(entry["name"]))
        ws.cell(row=i, column=2, value=entry["overall"])
        ws.cell(row=i, column=3, value=entry["rating"])
        ws.cell(row=i, column=4, value=entry["games"])
        ws.cell(row=i, column=5, value="ja" if entry["icon"] else "nee")
        ws.cell(row=i, column=6, value=given.get(person_id, 0))
        ws.cell(row=i, column=7, value=received.get(person_id, 0))
        ws.cell(row=i, column=8, value=owners.get(person_id, 0))

    for column, width in zip("ABCDEFGH", (16, 10, 10, 13, 8, 19, 16, 20)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(order) + 1}" if order else "A1:H1"


def main():
    folder = script_dir()
    input_path = os.path.join(folder, "Input.txt")

    if not os.path.isfile(input_path):
        print(f"FOUT: kan Input.txt niet vinden in: {folder}")
        print("Zet Genereer_Pack_Matrix_Rating.bat en genereer_pack_matrix_rating.py in")
        print("dezelfde map als Input.txt.")
        sys.exit(1)

    print(f"Input.txt gevonden: {input_path}")
    try:
        packs = load_packs(input_path)
    except ValueError as e:
        print(f"FOUT: {e}")
        sys.exit(1)
    print(f"{len(packs)} packs gelezen.")

    counts, people, stats, errors = parse_packs(packs)
    if errors:
        print(f"WAARSCHUWING: {len(errors)} pack(s) werden niet herkend en zijn overgeslagen.")

    if not people:
        print("Geen geldige packs gevonden, er is niets om te verwerken.")
        sys.exit(1)

    order = sort_people(people)
    wb = build_workbook(counts, people, order)
    build_summary_sheet(wb, counts, people)
    build_people_sheet(wb, counts, people, order)

    # Elke run schrijft een eigen bestand, dus ze horen bij elkaar in een eigen map
    # in plaats van los naast het script. outputs/ staat in .gitignore.
    output_dir = os.path.join(folder, "outputs")
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    out_name = f"Pack_Matrix_Rating_{timestamp}.xlsx"
    out_path = os.path.join(output_dir, out_name)
    wb.save(out_path)

    zonder_kaart = sum(1 for p in people.values() if p["overall"] is None)
    print(
        f"{len(people)} personen gevonden, {stats['cards']} kaarten "
        f"uit {stats['packs']} packs verwerkt."
    )
    if zonder_kaart:
        print(f"({zonder_kaart} daarvan hebben zelf geen kaart en staan onderaan de matrix.)")
    print(f"Klaar! Bestand opgeslagen als: outputs\\{out_name}")
    print(f"Volledig pad: {out_path}")


if __name__ == "__main__":
    main()
